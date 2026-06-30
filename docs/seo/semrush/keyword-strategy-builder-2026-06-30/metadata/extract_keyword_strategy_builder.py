#!/usr/bin/env python3
"""Extract Semrush Keyword Strategy Builder workbook data to CSV/JSONL."""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
ORIGINAL = ROOT / "originals" / "littleherolabs_com_-_keyword_strategy.xlsx"
CSV_OUT = ROOT / "csv" / "littleherolabs_com_-_keyword_strategy__table.csv"
JSONL_OUT = ROOT / "metadata" / "littleherolabs_com_-_keyword_strategy__table.cells.jsonl"
MANIFEST_OUT = ROOT / "metadata" / "manifest.json"
CHECKSUMS_OUT = ROOT / "metadata" / "checksums.sha256"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> None:
    workbook = openpyxl.load_workbook(ORIGINAL, data_only=True)
    worksheet = workbook["table"]

    rows = list(worksheet.iter_rows(values_only=True))
    header = [cell if cell is not None else "" for cell in rows[0]]
    data_rows = rows[1:]

    CSV_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSONL_OUT.parent.mkdir(parents=True, exist_ok=True)

    with CSV_OUT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        writer.writerows(data_rows)

    with JSONL_OUT.open("w", encoding="utf-8") as handle:
        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                handle.write(
                    json.dumps(
                        {
                            "sheet": "table",
                            "row": row_index,
                            "column": column_index,
                            "value": value,
                        },
                        ensure_ascii=False,
                    )
                    + "\n"
                )

    manifest = {
        "source": "Semrush Keyword Strategy Builder",
        "captured_at": "2026-06-30",
        "original_file": str(ORIGINAL.relative_to(ROOT)),
        "sheets": [
            {
                "name": sheet.title,
                "rows": sheet.max_row,
                "columns": sheet.max_column,
            }
            for sheet in workbook.worksheets
        ],
        "extracted_files": [
            str(CSV_OUT.relative_to(ROOT)),
            str(JSONL_OUT.relative_to(ROOT)),
        ],
        "table_rows_excluding_header": len(data_rows),
        "table_columns": header,
    }
    MANIFEST_OUT.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    checksums = [
        ORIGINAL,
        CSV_OUT,
        JSONL_OUT,
        MANIFEST_OUT,
    ]
    CHECKSUMS_OUT.write_text(
        "".join(f"{sha256(path)}  {path.relative_to(ROOT)}\n" for path in checksums),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
