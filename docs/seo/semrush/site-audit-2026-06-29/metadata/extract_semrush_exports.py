#!/usr/bin/env python3
"""Extract raw Semrush XLSX exports into CSV and metadata files.

This script is intentionally capture-only: it preserves every worksheet's
cell values without analysis, scoring, or recommendations.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


ARCHIVE_DIR = Path(__file__).resolve().parents[1]
ORIGINALS_DIR = ARCHIVE_DIR / "originals"
CSV_DIR = ARCHIVE_DIR / "csv"
METADATA_DIR = ARCHIVE_DIR / "metadata"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip()).strip("-")
    return slug or "sheet"


def json_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return value


def csv_value(cell: Cell) -> Any:
    value = json_value(cell.value)
    if value is None:
        return ""
    return value


def has_capture_content(cell: Cell) -> bool:
    return (
        cell.value is not None
        or cell.hyperlink is not None
        or cell.comment is not None
    )


def cell_metadata(cell: Cell) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "coordinate": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "value": json_value(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
    }
    if cell.hyperlink is not None:
        metadata["hyperlink"] = {
            "target": cell.hyperlink.target,
            "location": cell.hyperlink.location,
            "display": cell.hyperlink.display,
            "tooltip": cell.hyperlink.tooltip,
        }
    if cell.comment is not None:
        metadata["comment"] = {
            "author": cell.comment.author,
            "text": cell.comment.text,
        }
    return metadata


def extract_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    workbook_slug = safe_slug(path.stem)
    workbook_record: dict[str, Any] = {
        "file_name": path.name,
        "relative_path": str(path.relative_to(ARCHIVE_DIR)),
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "sheets": [],
    }

    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        sheet_slug = safe_slug(worksheet.title)
        prefix = f"{workbook_slug}__{sheet_index:02d}-{sheet_slug}"
        csv_path = CSV_DIR / f"{prefix}.csv"
        cells_path = METADATA_DIR / f"{prefix}.cells.jsonl"

        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        actual_max_row = 0
        actual_max_col = 0
        non_empty_cells = 0
        formula_cells = 0
        hyperlink_cells = 0
        comment_cells = 0

        with cells_path.open("w", encoding="utf-8") as cells_handle:
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_col,
            ):
                for cell in row:
                    if not has_capture_content(cell):
                        continue
                    non_empty_cells += 1
                    actual_max_row = max(actual_max_row, cell.row)
                    actual_max_col = max(actual_max_col, cell.column)
                    if cell.data_type == "f":
                        formula_cells += 1
                    if cell.hyperlink is not None:
                        hyperlink_cells += 1
                    if cell.comment is not None:
                        comment_cells += 1
                    cells_handle.write(json.dumps(cell_metadata(cell), ensure_ascii=False))
                    cells_handle.write("\n")

        with csv_path.open("w", encoding="utf-8", newline="") as csv_handle:
            writer = csv.writer(csv_handle, lineterminator="\n")
            if actual_max_row and actual_max_col:
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=actual_max_row,
                    min_col=1,
                    max_col=actual_max_col,
                ):
                    writer.writerow([csv_value(cell) for cell in row])

        workbook_record["sheets"].append(
            {
                "sheet_index": sheet_index,
                "sheet_name": worksheet.title,
                "xlsx_max_row": max_row,
                "xlsx_max_column": max_col,
                "captured_max_row": actual_max_row,
                "captured_max_column": actual_max_col,
                "non_empty_cells": non_empty_cells,
                "formula_cells": formula_cells,
                "hyperlink_cells": hyperlink_cells,
                "comment_cells": comment_cells,
                "csv_relative_path": str(csv_path.relative_to(ARCHIVE_DIR)),
                "csv_sha256": sha256_file(csv_path),
                "cells_jsonl_relative_path": str(cells_path.relative_to(ARCHIVE_DIR)),
                "cells_jsonl_sha256": sha256_file(cells_path),
            }
        )

    workbook.close()
    return workbook_record


def write_checksums(manifest: dict[str, Any]) -> None:
    paths: list[Path] = []
    for workbook in manifest["workbooks"]:
        paths.append(ARCHIVE_DIR / workbook["relative_path"])
        for sheet in workbook["sheets"]:
            paths.append(ARCHIVE_DIR / sheet["csv_relative_path"])
            paths.append(ARCHIVE_DIR / sheet["cells_jsonl_relative_path"])

    checksums_path = METADATA_DIR / "checksums.sha256"
    with checksums_path.open("w", encoding="utf-8") as handle:
        for path in sorted(paths):
            handle.write(f"{sha256_file(path)}  {path.relative_to(ARCHIVE_DIR)}\n")


def write_readme(manifest: dict[str, Any]) -> None:
    readme_path = ARCHIVE_DIR / "README.md"
    lines = [
        "# Semrush Site Audit Raw Export Capture - 2026-06-29",
        "",
        "This directory captures the Semrush site audit exports shared on 2026-06-29.",
        "It is an archive of raw source data, not an analysis or recommendation document.",
        "",
        "## Contents",
        "",
        "- `originals/`: exact `.xlsx` files copied from the shared Semrush exports.",
        "- `csv/`: one full-value CSV export per worksheet.",
        "- `metadata/manifest.json`: workbook, sheet, dimensions, and checksum metadata.",
        "- `metadata/*.cells.jsonl`: one JSON object per captured non-empty cell, including value, data type, number format, hyperlink, and comment fields when present.",
        "- `metadata/checksums.sha256`: SHA-256 checksums for captured originals and extracted files.",
        "",
        "## Source Files",
        "",
    ]
    for workbook in manifest["workbooks"]:
        lines.append(f"- `{workbook['relative_path']}`")
    lines.append("")
    readme_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    METADATA_DIR.mkdir(parents=True, exist_ok=True)

    manifest: dict[str, Any] = {
        "capture_name": "Semrush site audit raw export capture",
        "captured_at_local": dt.datetime.now().astimezone().isoformat(),
        "archive_date": "2026-06-29",
        "domain": "littleherolabs.com",
        "workbooks": [],
    }

    for path in sorted(ORIGINALS_DIR.glob("*.xlsx")):
        manifest["workbooks"].append(extract_workbook(path))

    manifest_path = METADATA_DIR / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    write_checksums(manifest)
    write_readme(manifest)


if __name__ == "__main__":
    main()
